import re
import io
import requests
from decimal import Decimal, InvalidOperation
from pathlib import Path
from django.core.management.base import BaseCommand
from django.core.files.base import ContentFile
from apps.catalog.models import Category, Product, ProductImage, Brand, Room

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36',
    'Accept-Language': 'ru-RU,ru;q=0.9',
}

CATEGORY_ALIASES = {
    'оборудование для контроля и защ': 'Оборудование для контроля и защиты',
}


def get_photo_url(product_url):
    try:
        r = requests.get(product_url, headers=HEADERS, timeout=15)
        m = re.search(r'og:image[^>]*content="([^"]+)"', r.text)
        if m:
            return m.group(1)
    except Exception:
        pass
    return None


def download_image(url):
    try:
        r = requests.get(url, headers=HEADERS, timeout=15)
        if r.status_code == 200:
            return r.content
    except Exception:
        pass
    return None


def parse_price(val):
    if val is None:
        return None
    s = str(val).replace('₽', '').replace(' ', '').replace('\xa0', '').strip()
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def parse_excel_file(filepath):
    import openpyxl
    wb = openpyxl.load_workbook(filepath)
    products = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        current_category = sheet_name.strip()

        # Detect header row
        header_row = None
        for row in range(1, ws.max_row + 1):
            vals = [ws.cell(row, c).value for c in range(1, ws.max_column + 1)]
            vals_str = [str(v).strip().lower() if v else '' for v in vals]
            if any(x in vals_str for x in ['ссылка', 'ссылка на товар']):
                header_row = row
                break

        if header_row is None:
            continue

        # Find column indices
        cols = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if v:
                v = str(v).strip().lower()
                if 'ссылка' in v:
                    cols['url'] = c
                elif 'название' in v and 'фото' not in v:
                    cols['name'] = c
                elif 'цена' in v:
                    cols['price'] = c
                elif 'бренд' in v:
                    cols['brand'] = c
                elif 'фото' in v:
                    cols['photo'] = c

        # Find category headers within sheet (for вторые товары format)
        # Scan ALL rows including before header
        cat_rows = {}
        for row in range(1, ws.max_row + 1):
            if row == header_row:
                continue
            first = ws.cell(row, 1).value
            if first and str(first).strip() and not any(
                ws.cell(row, c).value for c in range(2, ws.max_column + 1)
            ):
                cat_rows[row] = str(first).strip()

        # Set initial category: prefer last category header before header_row
        pre_header_cats = {r: v for r, v in cat_rows.items() if r < header_row}
        if pre_header_cats:
            current_cat_for_row = cat_rows[max(pre_header_cats.keys())]
        else:
            current_cat_for_row = current_category

        # Read products
        for row in range(header_row + 1, ws.max_row + 1):
            if row in cat_rows:
                current_cat_for_row = cat_rows[row]
                continue

            url = ws.cell(row, cols.get('url', 1)).value
            name = ws.cell(row, cols.get('name', 2)).value
            price = ws.cell(row, cols.get('price', 3)).value
            brand = ws.cell(row, cols.get('brand', 4)).value

            if not url or not name:
                continue
            url = str(url).strip()
            name = str(name).strip()
            if not url.startswith('http'):
                continue

            products.append({
                'url': url,
                'name': name,
                'price': parse_price(price),
                'brand': str(brand).strip() if brand else None,
                'category': current_cat_for_row,
            })

    return products


class Command(BaseCommand):
    help = 'Импорт товаров из Excel файлов с загрузкой фото с sanpalace.ru'

    def add_arguments(self, parser):
        parser.add_argument('files', nargs='+', help='Пути к Excel файлам')

    def handle(self, *args, **options):
        default_room = Room.objects.first()
        if not default_room:
            self.stdout.write(self.style.ERROR('Нет помещений в базе. Запусти create_sample_data сначала.'))
            return

        all_products = []
        for filepath in options['files']:
            self.stdout.write(f'Читаю {filepath}...')
            items = parse_excel_file(filepath)
            self.stdout.write(f'  Найдено: {len(items)} товаров')
            all_products.extend(items)

        self.stdout.write(f'\nВсего товаров: {len(all_products)}')
        self.stdout.write('='*60)

        created = 0
        skipped = 0
        no_photo = 0

        for item in all_products:
            name = item['name']
            price = item['price']
            category_name = item['category']
            brand_name = item['brand']
            product_url = item['url']

            # Skip if no price
            if not price:
                self.stdout.write(self.style.WARNING(f'  ⚠ Пропущен (нет цены): {name}'))
                skipped += 1
                continue

            # Skip if already exists
            if Product.objects.filter(name=name).exists():
                self.stdout.write(f'  — Уже есть: {name}')
                skipped += 1
                continue

            # Resolve category
            cat_lookup = CATEGORY_ALIASES.get(category_name.lower(), category_name)
            category = Category.objects.filter(name__iexact=cat_lookup).first()
            if not category:
                # Try partial match
                category = Category.objects.filter(name__icontains=cat_lookup[:10]).first()
            if not category:
                self.stdout.write(self.style.WARNING(f'  ⚠ Категория не найдена: {category_name}, пропуск'))
                skipped += 1
                continue

            # Resolve or create brand
            brand = None
            if brand_name:
                brand = Brand.objects.filter(name__iexact=brand_name).first()
                if not brand:
                    brand = Brand.objects.create(
                        name=brand_name,
                        description=f'Бренд {brand_name}',
                        is_active=True,
                    )
                    self.stdout.write(f'  + Создан бренд: {brand_name}')

            # Pick room
            room = category.rooms.first() or default_room

            # Generate SKU
            brand_prefix = (brand.slug[:3] if brand else 'PRD').upper()
            cat_prefix = category.slug[:3].upper()
            count = Product.objects.count() + 1
            sku = f'{brand_prefix}-{cat_prefix}-{1000 + count}'

            product = Product.objects.create(
                name=name,
                description=f'{name}. Категория: {category.name}.',
                category=category,
                room=room,
                brand=brand,
                price=price,
                stock_quantity=10,
                sku=sku,
                is_active=True,
            )

            # Download and attach photo
            self.stdout.write(f'  → Загружаю фото для: {name[:50]}...')
            photo_url = get_photo_url(product_url)
            if photo_url:
                img_data = download_image(photo_url)
                if img_data:
                    ext = photo_url.split('.')[-1].split('?')[0] or 'jpg'
                    filename = f'{product.slug}.{ext}'
                    pi = ProductImage(product=product, alt_text=name, is_main=True)
                    pi.image.save(filename, ContentFile(img_data), save=True)
                    self.stdout.write(self.style.SUCCESS(f'  ✓ {name[:50]} — фото OK'))
                else:
                    self.stdout.write(self.style.WARNING(f'  ⚠ {name[:50]} — не скачалось фото'))
                    no_photo += 1
            else:
                self.stdout.write(self.style.WARNING(f'  ⚠ {name[:50]} — фото не найдено'))
                no_photo += 1

            created += 1

        self.stdout.write('\n' + '='*60)
        self.stdout.write(self.style.SUCCESS(f'✓ Создано товаров: {created}'))
        self.stdout.write(f'  Без фото: {no_photo}')
        self.stdout.write(f'  Пропущено: {skipped}')
        self.stdout.write(f'  Всего в базе: {Product.objects.count()}')
