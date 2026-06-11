import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from django.core.management.base import BaseCommand
from django.core.files.base import ContentFile
from apps.catalog.models import Category, Product, ProductImage, Brand, Room

CATEGORY_ALIASES = {
    'оборудование для контроля и защ': 'Оборудование для контроля и защиты',
}


def parse_price(val):
    if val is None:
        return None
    s = str(val).replace('₽', '').replace(' ', '').replace('\xa0', '').strip()
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def find_photo(photo_dir, photo_name):
    if not photo_name or not photo_dir:
        return None
    photo_name = str(photo_name).strip()
    for ext in ['jpeg', 'jpg', 'png', 'webp']:
        p = photo_dir / f'{photo_name}.{ext}'
        if p.exists():
            return p
    # Fallback: search by prefix
    for f in photo_dir.iterdir():
        if f.stem.lower() == photo_name.lower():
            return f
    return None


def parse_excel_with_local_photos(filepath):
    import openpyxl
    wb = openpyxl.load_workbook(filepath)
    products = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        current_category = sheet_name.strip()

        # Detect header row
        header_row = None
        for row in range(1, ws.max_row + 1):
            vals = [ws.cell(row, c).value for c in range(1, ws.max_column + 1)]
            vals_str = [str(v).strip().lower() if v else '' for v in vals]
            if any(x in vals_str for x in ['ссылка', 'ссылка на товар', 'название']):
                header_row = row
                break

        if header_row is None:
            continue

        # Find column indices from header
        cols = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if v:
                v = str(v).strip().lower()
                if 'название' in v and 'фото' not in v:
                    cols['name'] = c
                elif 'цена' in v:
                    cols['price'] = c
                elif 'бренд' in v:
                    cols['brand'] = c
                elif 'фото' in v:
                    cols['photo'] = c

        # If no name column found, use col 2
        if 'name' not in cols:
            cols['name'] = 2

        # Find category headers (single-cell rows)
        cat_rows = {}
        for row in range(1, ws.max_row + 1):
            if row == header_row:
                continue
            first = ws.cell(row, 1).value
            if first and str(first).strip() and not any(
                ws.cell(row, c).value for c in range(2, ws.max_column + 1)
            ):
                cat_rows[row] = str(first).strip()

        # Set initial category from last cat_row before header
        pre_header_cats = {r: v for r, v in cat_rows.items() if r < header_row}
        if pre_header_cats:
            current_cat_for_row = cat_rows[max(pre_header_cats.keys())]
        else:
            current_cat_for_row = current_category

        for row in range(header_row + 1, ws.max_row + 1):
            if row in cat_rows:
                current_cat_for_row = cat_rows[row]
                continue

            name = ws.cell(row, cols.get('name', 2)).value
            price = ws.cell(row, cols.get('price', 3)).value
            photo = ws.cell(row, cols.get('photo', 4)).value
            brand = ws.cell(row, cols.get('brand', 5)).value

            if not name or not str(name).strip():
                continue

            products.append({
                'name': str(name).strip(),
                'price': parse_price(price),
                'photo_name': str(photo).strip() if photo else None,
                'brand': str(brand).strip() if brand else None,
                'category': current_cat_for_row,
            })

    return products


class Command(BaseCommand):
    help = 'Импорт товаров из Excel с локальными фото'

    def add_arguments(self, parser):
        parser.add_argument('excel', help='Путь к Excel файлу')
        parser.add_argument('photo_dir', help='Папка с фото')

    def handle(self, *args, **options):
        excel_path = options['excel']
        photo_dir = Path(options['photo_dir'])
        default_room = Room.objects.first()

        if not default_room:
            self.stdout.write(self.style.ERROR('Нет помещений. Запусти create_sample_data сначала.'))
            return

        if not photo_dir.exists():
            self.stdout.write(self.style.ERROR(f'Папка с фото не найдена: {photo_dir}'))
            return

        self.stdout.write(f'Читаю {excel_path}...')
        items = parse_excel_with_local_photos(excel_path)
        self.stdout.write(f'Найдено: {len(items)} товаров')
        self.stdout.write('='*60)

        created = 0
        skipped = 0
        no_photo = 0

        for item in items:
            name = item['name']
            price = item['price']
            category_name = item['category']
            brand_name = item['brand']
            photo_name = item['photo_name']

            if not price:
                self.stdout.write(self.style.WARNING(f'  ⚠ Нет цены: {name}'))
                skipped += 1
                continue

            if Product.objects.filter(name=name).exists():
                self.stdout.write(f'  — Уже есть: {name}')
                skipped += 1
                continue

            # Resolve category
            cat_lookup = CATEGORY_ALIASES.get(category_name.lower(), category_name)
            category = Category.objects.filter(name__iexact=cat_lookup).first()
            if not category:
                category = Category.objects.filter(name__icontains=cat_lookup[:10]).first()
            if not category:
                self.stdout.write(self.style.WARNING(f'  ⚠ Категория не найдена: {category_name}'))
                skipped += 1
                continue

            # Resolve or create brand
            brand = None
            if brand_name:
                brand = Brand.objects.filter(name__iexact=brand_name).first()
                if not brand:
                    brand = Brand.objects.create(name=brand_name, description=f'Бренд {brand_name}', is_active=True)
                    self.stdout.write(f'  + Создан бренд: {brand_name}')

            room = category.rooms.first() or default_room
            brand_prefix = (brand.slug[:3] if brand else 'PRD').upper()
            cat_prefix = category.slug[:3].upper()
            sku = f'{brand_prefix}-{cat_prefix}-{1000 + Product.objects.count() + 1}'

            product = Product.objects.create(
                name=name,
                description=f'{name}. Категория: {category.name}.',
                category=category,
                room=room,
                brand=brand,
                price=price,
                stock_quantity=10,
                sku=sku,
                is_active=True,
            )

            # Attach local photo
            photo_path = find_photo(photo_dir, photo_name)
            if photo_path:
                with open(photo_path, 'rb') as f:
                    img_data = f.read()
                ext = photo_path.suffix.lstrip('.')
                filename = f'{product.slug}.{ext}'
                pi = ProductImage(product=product, alt_text=name, is_main=True)
                pi.image.save(filename, ContentFile(img_data), save=True)
                self.stdout.write(self.style.SUCCESS(f'  ✓ {name[:55]} — фото OK'))
            else:
                self.stdout.write(self.style.WARNING(f'  ⚠ {name[:55]} — фото не найдено ({photo_name})'))
                no_photo += 1

            created += 1

        self.stdout.write('\n' + '='*60)
        self.stdout.write(self.style.SUCCESS(f'✓ Создано: {created}'))
        self.stdout.write(f'  Без фото: {no_photo}')
        self.stdout.write(f'  Пропущено: {skipped}')
        self.stdout.write(f'  Всего товаров в базе: {Product.objects.count()}')
